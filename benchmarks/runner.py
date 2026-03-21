"""Benchmark runner for comparing XML, JSON, and LION notation modes."""

import asyncio
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import tiktoken
from openai import AsyncOpenAI
from vertexai.preview import tokenization
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import config as app_config

try:
    from playwright.async_api import async_playwright
    from benchmarks.bpmn_executor import BpmnIoExecutor
except ImportError:
    sys.exit(
        "\nERROR: playwright is not installed — it is required to run the benchmark.\n"
        "Install with:\n"
        "    pip install playwright\n"
        "    playwright install chromium\n"
    )

from benchmarks.prompts import (
    BENCHMARK_XML_PROMPT_FINAL,
    BENCHMARK_JSON_PROMPT_FINAL,
    BENCHMARK_LION_PROMPT_FINAL,
)
from lion import loads as lion_loads, strip_markdown_fences
from benchmarks.config import (
    ACTION_ORDER,
    BENCHMARK_SETTINGS,
    LION_SCHEMA_MAPPINGS,
    MAX_WORKERS,
    NOTATION_MODES,
    REPETITIONS,
    RESULTS_DIR,
)

# Minimal valid BPMN XML returned when generation or parsing fails
_EMPTY_BPMN_XML = (
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<bpmn:definitions xmlns:bpmn="http://www.omg.org/spec/BPMN/20100524/MODEL" '
    'id="Definitions_1" targetNamespace="http://bpmn.io/schema/bpmn">'
    '</bpmn:definitions>'
)


def strip_line_comments(text: str) -> str:
    """Remove // line comments, preserving :// in URLs."""
    return re.sub(r'(?m)(?<![:/])//[^\n]*', '', text)


def _log(msg: str) -> None:
    """Print a timestamped message, flushed immediately."""
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"  [{ts}] {msg}", flush=True)


class TokenCounter:
    def __init__(self, model):
        try:
            self.encoding = tiktoken.encoding_for_model(model)
        except KeyError:
            self.encoding = tiktoken.get_encoding("cl100k_base")

    def count(self, text: str) -> int:
        if not text:
            return 0
        return len(self.encoding.encode(text))

    def count_messages(self, messages: List[Dict[str, str]]) -> int:
        total = 0
        for msg in messages:
            total += self.count(msg.get("content", ""))
            total += 1
        return total
    
class GeminiTokenCounter:
    def __init__(self):
        self.tokenizer = tokenization.get_tokenizer_for_model("gemini-1.5-pro-002")

    def count(self, text: str) -> int:
        if not text:
            return 0
        result = self.tokenizer.count_tokens(text)
        return result.total_tokens

    def count_messages(self, messages: List[Dict[str, str]]) -> int:
        total = 0
        for msg in messages:
            total += self.count(msg.get("content", ""))
            total += 1
        return total


class BenchmarkRunner:
    def __init__(self) -> None:
        self.model = app_config.GPT_MODEL
        self.token_counter = TokenCounter(self.model)
        self.gemini_counter = GeminiTokenCounter()
        self.tasks = app_config.TASKS
        self._log_dir: Optional[Path] = None
        self._completed = 0
        self._total_runs = 0


    def _init_log_dir(self, session_dir: Path) -> None:
        self._log_dir = session_dir / "logs"
        self._log_dir.mkdir(parents=True, exist_ok=True)

    def _log_interaction(
        self,
        task_id: str,
        notation: str,
        run_number: int,
        messages: List[Dict[str, str]],
        response: str,
    ) -> None:
        if self._log_dir is None:
            return
        log_file = self._log_dir / f"{task_id}_{notation}_run{run_number:02d}.md"
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # noqa: DTZ005
        parts: List[str] = [
            f"# LLM Log: {task_id} — {notation}",
            f"- **Model:** {self.model}",
            f"- **Time:** {timestamp}\n",
            "---\n",
            "### Input\n",
        ]
        for msg in messages:
            role = msg.get("role", "unknown").upper()
            content = msg.get("content", "")
            if len(content) > 8000:
                content = content[:500] + "\n\n... [truncated] ...\n\n" + content[-500:]
            parts.append(f"**[{role}]**\n```\n{content}\n```\n")
        parts += ["### Output\n", f"```\n{response}\n```\n", "---\n"]
        log_file.write_text("\n".join(parts), encoding="utf-8")


    def run(self) -> None:
        try:
            asyncio.run(self._run_async())
        except KeyboardInterrupt:
            print("\n  Benchmark interrupted by user (Ctrl+C).\n", flush=True)


    async def _run_async(self) -> None:
        session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_dir = RESULTS_DIR / session_id

        for notation in NOTATION_MODES:
            (session_dir / notation).mkdir(parents=True, exist_ok=True)
        self._init_log_dir(session_dir)

        jobs = [
            (task, notation, run_num)
            for task in self.tasks
            for notation in NOTATION_MODES
            for run_num in range(1, REPETITIONS + 1)
        ]
        self._total_runs = len(jobs)
        self._completed = 0
        results: Dict[str, List[Dict[str, Any]]] = {n: [] for n in NOTATION_MODES}

        semaphore = asyncio.Semaphore(MAX_WORKERS)
        client = AsyncOpenAI(api_key=os.getenv("OPENAI_API_KEY"))

        async def _run_job(job: tuple) -> tuple:
            task, notation, run_num = job
            label = f"{task['id']} | {notation} | Run {run_num}/{REPETITIONS}"
            async with semaphore:
                print(f"  [starting] {label} …", flush=True)
                try:
                    result = await asyncio.wait_for(
                        self._run_task(task, notation, run_num, session_dir, client, executor),
                        timeout=BENCHMARK_SETTINGS["timeout_seconds"],
                    )
                    self._completed += 1
                    print(
                        f"  [{self._completed}/{self._total_runs}] {label} – "
                        f"Time: {result['time_seconds']:.1f}s, "
                        f"Tokens (OpenAI): {result['total_tokens']}, "
                        f"Tokens (Gemini): {result.get('gemini_tokens', 0)}",
                        flush=True,
                    )
                    return notation, result
                except BaseException as exc:  # noqa: BLE001
                    self._completed += 1
                    err = f"{type(exc).__name__}: {exc}"
                    print(
                        f"  [{self._completed}/{self._total_runs}] {label} – ERROR: {err}",
                        flush=True,
                    )
                    return notation, {
                        "task_id": task["id"],
                        "task_title": task["title"],
                        "run_number": run_num,
                        "time_seconds": 0.0,
                        "total_tokens": 0,
                        "gemini_tokens": 0,
                        "accuracy": None,
                        "error": err,
                    }

        print(
            f"\nStarting {self._total_runs} runs "
            f"({len(self.tasks)} tasks × {len(NOTATION_MODES)} notations × "
            f"{REPETITIONS} repetitions) with max {MAX_WORKERS} concurrent tasks …\n"
        )

        async with async_playwright() as pw:
            executor = await BpmnIoExecutor.create(pw)
            _log("Headless bpmn.io executor ready (Playwright/Chromium).")
            try:
                job_results = await asyncio.gather(
                    *[_run_job(j) for j in jobs], return_exceptions=True
                )
            finally:
                await executor.close()

        for item in job_results:
            if isinstance(item, BaseException):
                _log(f"[Unhandled Error] {type(item).__name__}: {item}")
                continue
            notation, result = item
            results[notation].append(result)

        for notation in NOTATION_MODES:
            results[notation].sort(key=lambda d: (d["task_id"], d["run_number"]))

        self._generate_excel(session_dir, results)
        print(f"\n{'='*60}", flush=True)
        print(f"Benchmark complete. Results saved to: {session_dir}", flush=True)
        print(f"{'='*60}", flush=True)


    async def _run_task(
        self,
        task: Dict,
        notation: str,
        run_number: int,
        session_dir: Path,
        client: AsyncOpenAI,
        executor: "BpmnIoExecutor",
    ) -> Dict[str, Any]:
        label = f"{task['id']}|{notation}|R{run_number}"
        start = time.time()
        _log(f"  {label}: calling API …")

        if notation == "XML":
            bpmn_xml, total_tokens, gemini_tokens = await self._generate_xml(task, run_number, client)
        else:
            prompt = BENCHMARK_JSON_PROMPT_FINAL if notation == "JSON" else BENCHMARK_LION_PROMPT_FINAL
            bpmn_xml, total_tokens, gemini_tokens = await self._generate_action_based(
                task, notation, run_number, client, executor, prompt
            )

        elapsed = time.time() - start
        _log(f"  {label}: done ({elapsed:.1f}s), writing file …")

        bpmn_path = session_dir / notation / f"{task['id']}_run{run_number:02d}.bpmn"
        await asyncio.to_thread(bpmn_path.write_text, bpmn_xml, "utf-8")

        return {
            "task_id": task["id"],
            "task_title": task["title"],
            "run_number": run_number,
            "time_seconds": round(elapsed, 2),
            "total_tokens": total_tokens,
            "gemini_tokens": gemini_tokens,
            "accuracy": None,
        }


    async def _call_llm(
        self, messages: List[Dict[str, str]], client: AsyncOpenAI
    ) -> Tuple[str, int, int]:
        """Send *messages* to the model; return (content, openai_tokens, gemini_tokens)."""
        input_tokens = self.token_counter.count_messages(messages)
        response = await client.chat.completions.create(
            model=self.model, messages=messages
        )
        content = response.choices[0].message.content.strip()
        openai_tokens = input_tokens + self.token_counter.count(content)

        gemini_in = self.gemini_counter.count_messages(messages)
        gemini_out = self.gemini_counter.count(content)
        gemini_tokens = gemini_in + gemini_out

        return content, openai_tokens, gemini_tokens


    @staticmethod
    def _make_user_message(task: Dict) -> str:
        return (
            f"Create a complete BPMN model for the following process:\n\n"
            f"Title: {task['title']}\n\n"
            f"Description:\n{task['description']}\n\n"
            f"Generate the ENTIRE model in a single response."
        )


    async def _generate_xml(
        self, task: Dict, run_number: int, client: AsyncOpenAI
    ) -> Tuple[str, int, int]:
        label = f"{task['id']}|XML|R{run_number}"
        messages = [
            {"role": "system", "content": BENCHMARK_XML_PROMPT_FINAL},
            {"role": "user", "content": self._make_user_message(task)},
        ]
        content, total_tokens, gemini_tokens = await self._call_llm(messages, client)
        _log(f"  {label}: response received ({total_tokens} tokens OpenAI / {gemini_tokens} Gemini)")
        self._log_interaction(task["id"], "XML", run_number, messages, content)

        parsed = self._parse_json_response(content)
        bpmn_xml = self._clean_xml_response(
            parsed["bpmn_xml"] if parsed and "bpmn_xml" in parsed else content
        )

        if not bpmn_xml or "<bpmn:definitions" not in bpmn_xml:
            _log(f"  {label}: WARNING — no valid XML in response, using empty model")
            bpmn_xml = _EMPTY_BPMN_XML
        else:
            _log(f"  {label}: XML extracted successfully")

        return bpmn_xml, total_tokens, gemini_tokens


    async def _generate_action_based(
        self,
        task: Dict,
        notation: str,
        run_number: int,
        client: AsyncOpenAI,
        executor: "BpmnIoExecutor",
        system_prompt: str,
    ) -> Tuple[str, int, int]:
        """Shared generation path for JSON and LION notations."""
        label = f"{task['id']}|{notation}|R{run_number}"
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": self._make_user_message(task)},
        ]
        content, total_tokens, gemini_tokens = await self._call_llm(messages, client)
        _log(f"  {label}: response received ({total_tokens} tokens OpenAI / {gemini_tokens} Gemini)")
        self._log_interaction(task["id"], notation, run_number, messages, content)

        parsed = (
            self._parse_json_response(content)
            if notation == "JSON"
            else self._parse_lion_response(content)
        )
        if parsed is None:
            _log(f"  {label}: WARNING — parse failed, returning empty model")
            return _EMPTY_BPMN_XML, total_tokens, gemini_tokens

        actions = self._convert_actions(parsed.get("actions", {}))
        _log(f"  {label}: placing {len(actions)} actions via bpmn.io …")
        try:
            bpmn_xml = await executor.execute(actions)
        except Exception as exc:
            _log(f"  {label}: ERROR in bpmn.io executor — {exc}")
            bpmn_xml = _EMPTY_BPMN_XML

        return bpmn_xml, total_tokens, gemini_tokens


    def _clean_xml_response(self, content: str) -> str:
        content = strip_markdown_fences(content)
        if not content.startswith("<?xml"):
            idx = content.find("<?xml")
            if idx >= 0:
                content = content[idx:]
        return content


    def _parse_json_response(self, content: str) -> Optional[Dict]:
        content = strip_line_comments(strip_markdown_fences(content))
        try:
            return json.loads(content)
        except json.JSONDecodeError:
            preview = content[:200] + " …" if len(content) > 200 else content
            _log(f"  [JSON Parse Error] Could not parse response (preview: {preview})")
            return None


    def _parse_lion_response(self, content: str) -> Optional[Dict]:
        content = strip_line_comments(strip_markdown_fences(content))
        try:
            return lion_loads(content)
        except Exception as exc:
            preview = content[:200] + " …" if len(content) > 200 else content
            _log(f"  [LION Parse Error] {exc} (preview: {preview})")
            return None


    def _convert_actions(self, actions_raw: Dict) -> List[Dict[str, Any]]:
        """Convert a raw short-keyed actions dict to frontend-format action list."""
        if not isinstance(actions_raw, dict):
            return []

        actions: List[Dict[str, Any]] = []
        for action_type in ACTION_ORDER:
            items = actions_raw.get(action_type)
            if not isinstance(items, list):
                continue

            if action_type == 'delete':
                actions.extend(
                    {'type': 'delete', 'elementId': item}
                    for item in items if isinstance(item, str)
                )
                continue

            mapping = LION_SCHEMA_MAPPINGS.get(action_type, {})
            for item in items:
                if not isinstance(item, dict):
                    continue
                action: Dict[str, Any] = {'type': action_type}
                for short_key, frontend_key in mapping.items():
                    val = item.get(short_key)
                    if val is not None and val != '' and val != []:
                        action[frontend_key] = val
                if action_type == 'participate':
                    raw_lanes = action.get('lanes')
                    action['lanes'] = (
                        [str(lane) for lane in raw_lanes if lane]
                        if isinstance(raw_lanes, list)
                        else []
                    )
                actions.append(action)

        return actions
    

    def _generate_excel(self, session_dir: Path, results: Dict[str, List[Dict]]) -> None:
        wb = Workbook()
        ws_overview = wb.active
        ws_overview.title = "Overview"

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        center = Alignment(horizontal='center')

        def _header_row(ws, headers: List[str]) -> None:
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center

        # Overview sheet
        _header_row(ws_overview, ["Notation", "Accuracy", "Avg. Time (s)", "Avg. Tokens (OpenAI)", "Avg. Tokens (Gemini)"])
        for row_idx, notation in enumerate(NOTATION_MODES, 2):
            valid = [d for d in results.get(notation, []) if "error" not in d]
            n = max(len(valid), 1)
            row = [
                notation,
                None,  # Accuracy — filled in manually after evaluation
                round(sum(d["time_seconds"] for d in valid) / n, 2) if valid else 0,
                round(sum(d["total_tokens"] for d in valid) / n, 0) if valid else 0,
                round(sum(d.get("gemini_tokens", 0) for d in valid) / n, 0) if valid else 0,
            ]
            for col, val in enumerate(row, 1):
                cell = ws_overview.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.alignment = center
        _auto_width(ws_overview)

        # Per-notation detail sheets
        detail_headers = ["Task ID", "Task Title", "Run #", "Accuracy", "Time (s)", "Tokens (OpenAI)", "Tokens (Gemini)", "Error"]
        for notation in NOTATION_MODES:
            ws = wb.create_sheet(title=notation)
            _header_row(ws, detail_headers)
            for row_idx, d in enumerate(results.get(notation, []), 2):
                row = [
                    d.get("task_id", ""), d.get("task_title", ""), d.get("run_number", ""),
                    d.get("accuracy"), d.get("time_seconds", 0),
                    d.get("total_tokens", 0), d.get("gemini_tokens", 0), d.get("error", ""),
                ]
                for col, val in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.border = thin_border
                    if col >= 3:
                        cell.alignment = center
            _auto_width(ws)

        excel_path = session_dir / "results.xlsx"
        wb.save(str(excel_path))
        _log(f"Excel report saved: {excel_path}")


def _auto_width(ws) -> None:
    """Auto-fit column widths in an openpyxl worksheet to its content."""
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = width + 4


def main() -> None:
    runner = BenchmarkRunner()
    print("=" * 60)
    print("BPMN Benchmark – XML vs JSON vs LION")
    print(f"Model:         {app_config.GPT_MODEL}")
    print(f"Tasks:         {len(app_config.TASKS)}")
    print(f"Repetitions:   {REPETITIONS}")
    print(f"Concurrency:   {MAX_WORKERS}")
    print("=" * 60)
    runner.run()


if __name__ == "__main__":
    main()
